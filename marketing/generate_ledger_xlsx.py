"""
Hugonomy Marketing Experiment Ledger — .xlsx generator
Run: python generate_ledger_xlsx.py
Output: marketing_experiments_master.xlsx (same folder)

Requires: pip install openpyxl
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

INPUT_CSV  = "marketing_experiments_master.csv"
OUTPUT_XLSX = "marketing_experiments_master.xlsx"

PLATFORM_COLORS = {
    "TikTok":    "FF0050",
    "LinkedIn":  "0A66C2",
    "Facebook":  "1877F2",
    "MailerLite":"14B389",
}

DECISION_COLORS = {
    "Scale":       "00B050",
    "Repeat":      "92D050",
    "Remix":       "FFEB9C",
    "Investigate": "BDD7EE",
    "Kill":        "FFC7CE",
}

HOOK_STRUCTURE_COLORS = {
    "Threat":       "FFD7D7",
    "Announcement": "FFF3CD",
    "Concept":      "D7E8FF",
}

HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="F0C040", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
PIVOT_HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")

thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def safe_float(v):
    try:
        return float(str(v).replace("%", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def build_ledger_sheet(wb, rows):
    headers = rows[0]
    data_rows = rows[1:]

    ws = wb.active
    ws.title = "Experiment Ledger"

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.row_dimensions[1].height = 36

    platform_col       = headers.index("Platform")       if "Platform"       in headers else None
    decision_col       = headers.index("Decision")       if "Decision"       in headers else None
    hook_struct_col    = headers.index("HookStructure")  if "HookStructure"  in headers else None
    conv_conf_col      = headers.index("ConversionConfidence") if "ConversionConfidence" in headers else None

    for row_idx, row in enumerate(data_rows, start=2):
        is_alt = (row_idx % 2 == 0)
        # Pad short rows to avoid IndexError
        row = list(row) + [""] * (len(headers) - len(row))

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER

            zero_based = col_idx - 1

            if zero_based == 0 and platform_col is not None:
                hex_c = PLATFORM_COLORS.get(row[platform_col])
                if hex_c:
                    cell.fill = cell_fill(hex_c)
                    cell.font = Font(color="FFFFFF", bold=True, size=10)
                elif is_alt:
                    cell.fill = ALT_FILL

            elif hook_struct_col is not None and zero_based == hook_struct_col:
                hex_c = HOOK_STRUCTURE_COLORS.get(value)
                if hex_c:
                    cell.fill = cell_fill(hex_c)
                    cell.font = Font(bold=True, size=9)
                elif is_alt:
                    cell.fill = ALT_FILL

            elif decision_col is not None and zero_based == decision_col and value:
                key = value.split("—")[0].strip().split(" ")[0]
                hex_c = DECISION_COLORS.get(key)
                if hex_c:
                    cell.fill = cell_fill(hex_c)
                    cell.font = Font(bold=True, size=9)
                elif is_alt:
                    cell.fill = ALT_FILL

            elif conv_conf_col is not None and zero_based == conv_conf_col:
                if value == "Instrumented":
                    cell.fill = cell_fill("D4EDDA")
                    cell.font = Font(bold=True, size=9, color="155724")
                elif value == "Reach-only":
                    cell.fill = cell_fill("F8D7DA")
                    cell.font = Font(size=9, color="721C24")
                elif is_alt:
                    cell.fill = ALT_FILL

            elif is_alt:
                cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 80

    COL_WIDTHS = {
        "ExperimentID":          10,
        "Date":                  14,
        "Platform":              12,
        "ChannelType":           12,
        "ContentTitle":          28,
        "ContentURL":            22,
        "SourceTag":             20,
        "HookType":              18,
        "HookStructure":         14,
        "PrimaryGoal":           18,
        "ConversionConfidence":  18,
        "Hook":                  38,
        "CoreMessage":           38,
        "CTA":                   20,
        "TargetAudience":        22,
        "Views":                  9,
        "Impressions":           12,
        "WatchTime":             12,
        "RetentionRate":         13,
        "Likes":                  8,
        "Comments":               9,
        "Shares":                 8,
        "Saves":                  8,
        "ProfileVisits":         13,
        "WebsiteClicks":         14,
        "ExtensionInstalls":     16,
        "WaitlistSignups":       15,
        "Notes_Human":           45,
        "Notes_Gemini":          45,
        "Notes_Claude":          45,
        "Decision":              45,
        "DecisionStatus":        14,
    }

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 20)

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions

    return headers, data_rows


def build_pivot_sheet(wb, headers, data_rows):
    """
    Pivot tab: HookStructure × SourceTag
    Shows avg Views, avg RetentionRate, total WaitlistSignups per group.
    Excludes Account Summary rows (HookStructure = N/A).
    Answers: which hook structure and which channel actually produce leads?
    """
    ws = wb.create_sheet("Pivot — Hook × Channel")

    def col(name):
        return headers.index(name) if name in headers else None

    views_c      = col("Views")
    retention_c  = col("RetentionRate")
    signups_c    = col("WaitlistSignups")
    hook_c       = col("HookStructure")
    source_c     = col("SourceTag")
    conv_conf_c  = col("ConversionConfidence")

    # Aggregate: key = (HookStructure, SourceTag or "untagged")
    groups = defaultdict(lambda: {"views": [], "retention": [], "signups": 0, "rows": 0})

    for row in data_rows:
        row = list(row) + [""] * (len(headers) - len(row))
        hook_struct = row[hook_c] if hook_c is not None else ""
        if not hook_struct or hook_struct in ("N/A", ""):
            continue

        source = row[source_c] if source_c is not None else ""
        source = source.strip() if source else "untagged"

        key = (hook_struct, source)
        groups[key]["rows"] += 1

        v = safe_float(row[views_c]) if views_c is not None else None
        if v is not None:
            groups[key]["views"].append(v)

        r = safe_float(row[retention_c]) if retention_c is not None else None
        if r is not None:
            groups[key]["retention"].append(r)

        s = safe_float(row[signups_c]) if signups_c is not None else None
        if s is not None:
            groups[key]["signups"] += s

    # Header
    pivot_headers = ["HookStructure", "SourceTag", "Rows", "Avg Views", "Avg RetentionRate (%)", "Total WaitlistSignups"]
    for col_idx, h in enumerate(pivot_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = PIVOT_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 36

    # Sort: Threat first, then Announcement, then Concept; within group sort by avg views desc
    order = {"Threat": 0, "Announcement": 1, "Concept": 2}
    sorted_keys = sorted(groups.keys(), key=lambda k: (order.get(k[0], 99),
                         -sum(groups[k]["views"]) / max(len(groups[k]["views"]), 1)))

    for row_idx, key in enumerate(sorted_keys, start=2):
        hook_struct, source = key
        g = groups[key]
        avg_views     = round(sum(g["views"]) / len(g["views"]), 1) if g["views"] else ""
        avg_retention = round(sum(g["retention"]) / len(g["retention"]), 1) if g["retention"] else ""
        signups       = int(g["signups"]) if g["signups"] else 0

        row_vals = [hook_struct, source, g["rows"], avg_views, avg_retention, signups]
        is_alt = (row_idx % 2 == 0)

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER

            if col_idx == 1:
                hex_c = HOOK_STRUCTURE_COLORS.get(hook_struct)
                if hex_c:
                    cell.fill = cell_fill(hex_c)
                    cell.font = Font(bold=True, size=10)
                elif is_alt:
                    cell.fill = ALT_FILL
            elif col_idx == 6 and val and val > 0:
                cell.fill = cell_fill("00B050")
                cell.font = Font(bold=True, color="FFFFFF")
            elif is_alt:
                cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 24

    pivot_col_widths = [16, 24, 8, 14, 22, 22]
    for i, w in enumerate(pivot_col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Note below pivot
    note_row = len(sorted_keys) + 3
    ws.cell(row=note_row, column=1, value="Note: pivot excludes Account Summary rows (HookStructure = N/A). Regenerate after each CSV update. Never hand-edit this sheet.")
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="888888")


def build_xlsx(csv_path, xlsx_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    if not rows:
        print("Empty CSV — nothing to write.")
        return

    headers, data_rows = build_ledger_sheet(openpyxl.Workbook(), rows)

    # Rebuild workbook so active sheet is set correctly
    wb = openpyxl.Workbook()
    headers, data_rows = build_ledger_sheet.__wrapped__(wb, rows) if hasattr(build_ledger_sheet, "__wrapped__") else _build_both(wb, rows)

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")
    print(f"  {len(data_rows)} data rows, {len(headers)} columns")


def _build_both(wb, rows):
    headers, data_rows = build_ledger_sheet(wb, rows)
    build_pivot_sheet(wb, headers, data_rows)
    return headers, data_rows


if __name__ == "__main__":
    with open(INPUT_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    if not rows:
        print("Empty CSV.")
    else:
        wb = openpyxl.Workbook()
        headers, data_rows = build_ledger_sheet(wb, rows)
        build_pivot_sheet(wb, headers, data_rows)
        wb.save(OUTPUT_XLSX)
        print(f"Saved: {OUTPUT_XLSX}")
        print(f"  {len(data_rows)} data rows, {len(headers)} columns, 2 sheets")
